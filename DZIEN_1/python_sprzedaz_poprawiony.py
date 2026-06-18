
"""
Analiza sprzedaży z pliku CSV - wersja na publicznych bibliotekach.

Wymagane biblioteki:
    pip install pandas openpyxl matplotlib

Skrypt:
1. wczytuje plik CSV:
   - separator: ;
   - kodowanie: utf-8-sig
   - separator dziesiętny: przecinek
2. liczy:
   - sprzedaż według miesiąca,
   - sprzedaż według kategorii,
   - sprzedaż według miasta,
   - top 10 produktów,
   - średni rabat,
3. zapisuje wyniki do pliku XLSX,
4. dodaje 3 wykresy w osobnych arkuszach Excela.

Autor: wersja edukacyjna
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# KONFIGURACJA
# ============================================================

INPUT_CSV = Path("sprzedaz_poprawne.csv")
OUTPUT_XLSX = Path("analiza_sprzedazy_publiczne_biblioteki.xlsx")

CSV_SEPARATOR = ";"
CSV_ENCODING = "utf-8-sig"
DECIMAL_SEPARATOR = ","

# Reguła biznesowa:
# do sprzedaży wliczamy tylko zamówienia zrealizowane.
STATUS_SPRZEDAZY = "Zrealizowane"


# ============================================================
# FUNKCJE POMOCNICZE
# ============================================================

def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Usuwa nadmiarowe spacje z nazw kolumn.
    """
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    return df


def prepare_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Przygotowuje typy danych po wczytaniu CSV.
    """
    df = df.copy()

    # Data
    df["data_zamowienia"] = pd.to_datetime(
        df["data_zamowienia"],
        errors="coerce"
    )

    # Kolumny liczbowe.
    # read_csv(decimal=",") powinno je odczytać poprawnie,
    # ale ta część zabezpiecza skrypt, gdyby coś zostało tekstem.
    numeric_columns = ["ilosc", "cena_jedn", "rabat_proc", "wartosc"]

    for col in numeric_columns:
        if col in df.columns:
            if df[col].dtype == "object":
                df[col] = (
                    df[col]
                    .astype(str)
                    .str.replace(" ", "", regex=False)
                    .str.replace(",", ".", regex=False)
                )
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Miesiąc jako tekst YYYY-MM
    df["miesiac"] = df["data_zamowienia"].dt.to_period("M").astype(str)

    return df


def filter_sales(df: pd.DataFrame) -> pd.DataFrame:
    """
    Zwraca tylko zamówienia zrealizowane.
    """
    return df[df["status"] == STATUS_SPRZEDAZY].copy()


def autosize_columns(ws):
    """
    Dopasowuje szerokość kolumn do treści.
    """
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        width = min(max(max_length + 2, 10), 40)
        ws.column_dimensions[column_letter].width = width


def style_sheet(ws):
    """
    Podstawowe formatowanie arkusza:
    - ciemny nagłówek,
    - biała czcionka,
    - zamrożony pierwszy wiersz,
    - obramowania.
    """
    header_fill = PatternFill(
        start_color="0F172A",
        end_color="0F172A",
        fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    autosize_columns(ws)


def format_numbers(ws):
    """
    Formatuje liczby w arkuszach.
    """
    money_headers = {
        "Sprzedaż",
        "Średnia wartość zamówienia",
        "Sprzedaż razem"
    }

    integer_headers = {
        "Liczba zamówień",
        "Ilość sprzedana",
        "Liczba rekordów",
        "Liczba zamówień zrealizowanych"
    }

    percent_headers = {
        "Średni rabat [%]"
    }

    header_to_col = {}

    for cell in ws[1]:
        header_to_col[cell.value] = cell.column

    for header, col_idx in header_to_col.items():
        col_letter = get_column_letter(col_idx)

        if header in money_headers:
            for cell in ws[col_letter][1:]:
                cell.number_format = '# ##0,00 zł'

        elif header in integer_headers:
            for cell in ws[col_letter][1:]:
                cell.number_format = '0'

        elif header in percent_headers:
            for cell in ws[col_letter][1:]:
                cell.number_format = '0,00'


def add_line_chart(ws, title, min_row, max_row, category_col, value_col, position):
    """
    Dodaje wykres liniowy do arkusza.
    """
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Sprzedaż"
    chart.x_axis.title = "Miesiąc"
    chart.height = 12
    chart.width = 22

    data = Reference(
        ws,
        min_col=value_col,
        min_row=min_row,
        max_row=max_row
    )
    categories = Reference(
        ws,
        min_col=category_col,
        min_row=min_row + 1,
        max_row=max_row
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, position)


def add_bar_chart(ws, title, min_row, max_row, category_col, value_col, position):
    """
    Dodaje wykres słupkowy do arkusza.
    """
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Sprzedaż"
    chart.x_axis.title = "Kategoria"
    chart.height = 12
    chart.width = 22

    data = Reference(
        ws,
        min_col=value_col,
        min_row=min_row,
        max_row=max_row
    )
    categories = Reference(
        ws,
        min_col=category_col,
        min_row=min_row + 1,
        max_row=max_row
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, position)


def write_report_to_excel(
    output_xlsx: Path,
    readme_df: pd.DataFrame,
    sales_by_month: pd.DataFrame,
    sales_by_category: pd.DataFrame,
    sales_by_city: pd.DataFrame,
    top10_products: pd.DataFrame,
    avg_discount: pd.DataFrame
):
    """
    Zapisuje raport do XLSX i dodaje formatowanie oraz wykresy.
    """
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="README", index=False)
        sales_by_month.to_excel(writer, sheet_name="Sprzedaz_miesiac", index=False)
        sales_by_category.to_excel(writer, sheet_name="Sprzedaz_kategoria", index=False)
        sales_by_city.to_excel(writer, sheet_name="Sprzedaz_miasto", index=False)
        top10_products.to_excel(writer, sheet_name="Top10_produkty", index=False)
        avg_discount.to_excel(writer, sheet_name="Sredni_rabat", index=False)

    wb = load_workbook(output_xlsx)

    for ws in wb.worksheets:
        style_sheet(ws)
        format_numbers(ws)

    # Wykres 1: sprzedaż według miesiąca
    ws_month = wb["Sprzedaz_miesiac"]
    add_line_chart(
        ws=ws_month,
        title="Sprzedaż według miesiąca",
        min_row=1,
        max_row=ws_month.max_row,
        category_col=1,
        value_col=2,
        position="F2"
    )

    # Wykres 2: sprzedaż według kategorii
    ws_category = wb["Sprzedaz_kategoria"]
    add_bar_chart(
        ws=ws_category,
        title="Sprzedaż według kategorii",
        min_row=1,
        max_row=ws_category.max_row,
        category_col=1,
        value_col=2,
        position="F2"
    )

    # Wykres 3: Top 10 produktów
    ws_products = wb["Top10_produkty"]
    add_bar_chart(
        ws=ws_products,
        title="Top 10 produktów według sprzedaży",
        min_row=1,
        max_row=ws_products.max_row,
        category_col=1,
        value_col=2,
        position="G2"
    )

    wb.save(output_xlsx)


# ============================================================
# GŁÓWNA ANALIZA
# ============================================================

def main():
    # --------------------------------------------------------
    # 1. Wczytanie CSV
    # --------------------------------------------------------

    df = pd.read_csv(
        INPUT_CSV,
        sep=CSV_SEPARATOR,
        encoding=CSV_ENCODING,
        decimal=DECIMAL_SEPARATOR
    )

    df = normalize_columns(df)
    df = prepare_data(df)

    sales_df = filter_sales(df)

    # --------------------------------------------------------
    # 2. Sprzedaż według miesiąca
    # --------------------------------------------------------

    sales_by_month = (
        sales_df
        .groupby("miesiac", as_index=False)
        .agg(
            Sprzedaż=("wartosc", "sum"),
            **{"Liczba zamówień": ("nr_zamowienia", "count")},
            **{"Średnia wartość zamówienia": ("wartosc", "mean")}
        )
        .rename(columns={"miesiac": "Miesiąc"})
        .sort_values("Miesiąc")
    )

    # --------------------------------------------------------
    # 3. Sprzedaż według kategorii
    # --------------------------------------------------------

    sales_by_category = (
        sales_df
        .groupby("kategoria", as_index=False)
        .agg(
            Sprzedaż=("wartosc", "sum"),
            **{"Liczba zamówień": ("nr_zamowienia", "count")},
            **{"Średnia wartość zamówienia": ("wartosc", "mean")}
        )
        .rename(columns={"kategoria": "Kategoria"})
        .sort_values("Sprzedaż", ascending=False)
    )

    # --------------------------------------------------------
    # 4. Sprzedaż według miasta
    # --------------------------------------------------------

    sales_by_city = (
        sales_df
        .groupby("miasto", as_index=False)
        .agg(
            Sprzedaż=("wartosc", "sum"),
            **{"Liczba zamówień": ("nr_zamowienia", "count")},
            **{"Średnia wartość zamówienia": ("wartosc", "mean")}
        )
        .rename(columns={"miasto": "Miasto"})
        .sort_values("Sprzedaż", ascending=False)
    )

    # --------------------------------------------------------
    # 5. Top 10 produktów
    # --------------------------------------------------------

    top10_products = (
        sales_df
        .groupby("produkt", as_index=False)
        .agg(
            Sprzedaż=("wartosc", "sum"),
            **{"Liczba zamówień": ("nr_zamowienia", "count")},
            **{"Ilość sprzedana": ("ilosc", "sum")},
            **{"Średnia wartość zamówienia": ("wartosc", "mean")}
        )
        .rename(columns={"produkt": "Produkt"})
        .sort_values("Sprzedaż", ascending=False)
        .head(10)
    )

    # --------------------------------------------------------
    # 6. Średni rabat
    # --------------------------------------------------------

    avg_discount_total = pd.DataFrame({
        "Analiza": ["Średni rabat ogółem"],
        "Średni rabat [%]": [sales_df["rabat_proc"].mean()],
        "Liczba zamówień": [sales_df["nr_zamowienia"].count()]
    })

    avg_discount_by_category = (
        sales_df
        .groupby("kategoria", as_index=False)
        .agg(
            **{"Średni rabat [%]": ("rabat_proc", "mean")},
            **{"Liczba zamówień": ("nr_zamowienia", "count")}
        )
        .rename(columns={"kategoria": "Analiza"})
    )

    avg_discount_by_category["Analiza"] = (
        "Średni rabat: " + avg_discount_by_category["Analiza"].astype(str)
    )

    avg_discount = pd.concat(
        [avg_discount_total, avg_discount_by_category],
        ignore_index=True
    )

    # --------------------------------------------------------
    # 7. README / podsumowanie
    # --------------------------------------------------------

    readme_df = pd.DataFrame({
        "Parametr": [
            "Plik źródłowy",
            "Separator CSV",
            "Kodowanie",
            "Separator dziesiętny",
            "Reguła sprzedaży",
            "Liczba rekordów w pliku",
            "Liczba zamówień zrealizowanych",
            "Sprzedaż razem",
            "Średni rabat",
            "Zakres dat od",
            "Zakres dat do"
        ],
        "Wartość": [
            str(INPUT_CSV),
            CSV_SEPARATOR,
            CSV_ENCODING,
            DECIMAL_SEPARATOR,
            f"Tylko status: {STATUS_SPRZEDAZY}",
            len(df),
            len(sales_df),
            round(sales_df["wartosc"].sum(), 2),
            round(sales_df["rabat_proc"].mean(), 2),
            str(sales_df["data_zamowienia"].min().date()),
            str(sales_df["data_zamowienia"].max().date())
        ]
    })

    # --------------------------------------------------------
    # 8. Zapis XLSX
    # --------------------------------------------------------

    write_report_to_excel(
        output_xlsx=OUTPUT_XLSX,
        readme_df=readme_df,
        sales_by_month=sales_by_month,
        sales_by_category=sales_by_category,
        sales_by_city=sales_by_city,
        top10_products=top10_products,
        avg_discount=avg_discount
    )

    # --------------------------------------------------------
    # 9. Komunikat końcowy
    # --------------------------------------------------------

    print("Analiza zakończona.")
    print(f"Plik XLSX zapisany jako: {OUTPUT_XLSX}")
    print(f"Liczba rekordów: {len(df)}")
    print(f"Liczba zamówień zrealizowanych: {len(sales_df)}")
    print(f"Sprzedaż razem: {sales_df['wartosc'].sum():.2f}")
    print(f"Średni rabat: {sales_df['rabat_proc'].mean():.2f}%")


if __name__ == "__main__":
    main()
